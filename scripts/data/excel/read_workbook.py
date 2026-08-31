"""Read-only Excel evidence extractor for the cross-validation pipeline.

This module intentionally exposes no save path. It opens the workbook twice in
read-only mode so formula text and cached values can be reported separately.
"""

from __future__ import annotations

import hashlib
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EXPECTED_SIZE = 3_206_663
EXPECTED_SHA256 = "7efe9af08bc11b5f6f28e006da3cc34db9ec637b11732934f71988ad6d553156"
EXPECTED_MTIME_NS = 1_788_187_104_562_882_000

RANGES: dict[str, tuple[int, int, int, int]] = {
    "属性克制": (1, 80, 1, 36),
    "全国图鉴": (1, 1293, 1, 21),
    "全形态图鉴": (1, 1447, 1, 19),
    "等级": (1, 264, 1, 36),
    "能力值": (1, 29, 1, 21),
    "属性PK": (1, 172, 1, 31),
    "特性列表": (1, 320, 1, 8),
    "招式列表": (1, 953, 1, 9),
    "猜宝可梦": (1, 1216, 1, 26),
    "关都图鉴": (1, 154, 1, 8),
    "伽勒尔图鉴": (1, 903, 1, 14),
    "铠岛图鉴": (1, 213, 1, 8),
    "雪原图鉴": (1, 211, 1, 8),
    "神奥图鉴": (1, 166, 1, 9),
    "洗翠图鉴": (1, 246, 1, 119),
    "帕底亚图鉴": (1, 937, 1, 14),
    "北上乡图鉴": (1, 201, 1, 11),
    "蓝莓学园图鉴": (1, 244, 1, 12),
    "密阿雷图鉴": (1, 233, 1, 10),
    "超次元图鉴": (1, 156, 1, 8),
    "Mega进化": (1, 94, 1, 15),
}


def fingerprint(path: Path) -> dict[str, Any]:
    stat = path.stat()
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return {
        "relativePath": "data-source/Pokemon-data.xlsx",
        "size": stat.st_size,
        "sha256": digest.hexdigest(),
        "mtimeUtc": f"{datetime.fromtimestamp(stat.st_mtime_ns // 1_000_000_000, timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')}.{stat.st_mtime_ns % 1_000_000_000 // 100:07d}Z",
        "mtimeNs": stat.st_mtime_ns,
    }


def assert_expected_fingerprint(value: dict[str, Any]) -> None:
    if value["size"] == EXPECTED_SIZE and value["sha256"] == EXPECTED_SHA256 and value["mtimeNs"] == EXPECTED_MTIME_NS:
        return
    raise RuntimeError(f"EXCEL_FINGERPRINT_MISMATCH: expected fixed {EXPECTED_SIZE}/{EXPECTED_SHA256}/{EXPECTED_MTIME_NS}, received {value['size']}/{value['sha256']}/{value['mtimeNs']}")


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "text"):
        return str(value.text)
    return str(value)


def extract_sheet(formula_sheet: Any, cached_sheet: Any, bounds: tuple[int, int, int, int]) -> dict[str, Any]:
    min_row, max_row, min_col, max_col = bounds
    cells: list[dict[str, Any]] = []
    formula_rows = formula_sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    )
    cached_rows = cached_sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    )
    for formula_row, cached_row in zip(formula_rows, cached_rows):
        for formula_cell, cached_cell in zip(formula_row, cached_row):
            raw = json_value(formula_cell.value)
            cached = json_value(cached_cell.value)
            if raw is None and cached is None:
                continue
            is_formula = formula_cell.data_type == "f" or type(formula_cell.value).__name__ == "ArrayFormula"
            cells.append({
                "row": formula_cell.row,
                "column": formula_cell.column,
                "locator": f"{formula_sheet.title}!{get_column_letter(formula_cell.column)}{formula_cell.row}",
                "kind": "formula" if is_formula else "static",
                "raw": raw,
                "cached": cached if is_formula else raw,
            })
    return {
        "title": formula_sheet.title,
        "state": formula_sheet.sheet_state,
        "maxRow": formula_sheet.max_row,
        "maxColumn": formula_sheet.max_column,
        "cells": cells,
    }


def main() -> None:
    if len(sys.argv) != 2:
        raise RuntimeError("Usage: read_workbook.py <workbook-path>")
    path = Path(sys.argv[1]).resolve()
    before = fingerprint(path)
    assert_expected_fingerprint(before)
    formula_book = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    cached_book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        missing = sorted(set(RANGES) - set(formula_book.sheetnames))
        if missing:
            raise RuntimeError(f"EXCEL_SHEETS_MISSING: {missing}")
        sheets = {
            title: extract_sheet(formula_book[title], cached_book[title], bounds)
            for title, bounds in RANGES.items()
        }
    finally:
        formula_book.close()
        cached_book.close()
    after = fingerprint(path)
    assert_expected_fingerprint(after)
    if before != after:
        raise RuntimeError("EXCEL_FINGERPRINT_CHANGED_DURING_READ")
    json.dump({
        "schemaVersion": 1,
        "adapter": "ExcelValidationAdapter",
        "readOnly": True,
        "saveCapability": False,
        "openpyxlVersion": __import__("openpyxl").__version__,
        "fingerprint": before,
        "sheets": sheets,
    }, sys.stdout, ensure_ascii=False, separators=(",", ":"))


if __name__ == "__main__":
    main()
