"""Materialize the reviewed 52Poké Item-name snapshot from the local workbook.

The workbook is an input container only.  This script reads the two explicitly
added Item sheets and never opens the workbook for writing.  The generated JSON
is the auditable build input; runtime generation does not read Excel.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_WORKBOOK_RELATIVE_PATH = "data-source/52poke-item-localization.xlsx"
SOURCE_WORKBOOK_PATH = PROJECT_ROOT / SOURCE_WORKBOOK_RELATIVE_PATH
SOURCE_WORKBOOK_SIZE = 3_533_083
SOURCE_WORKBOOK_SHA256 = "1731b09471fd92b070b37b9feb6c22bd80c1efa48e52a7b42da32fd453ebb3a7"
SOURCE_NAME = "神奇宝贝百科 / 52Poké Wiki"
SOURCE_URL = "https://wiki.52poke.com/wiki/道具列表（在其他语言中）"
LICENSE_URL = "https://wiki.52poke.com/wiki/神奇宝贝百科:版权声明"
COPIED_AT = "2026-08-31"

OWNER_OVERRIDES: dict[str, dict[str, str]] = {
    "snowball": {"zhHansName": "雪球", "reason": "主表同一英文键存在“雪球”和“雪丸”两条人工整理记录，按 owner 确认保留“雪球”。"},
    "mail": {"zhHansName": "邮件", "reason": "主表没有通用 Mail 行；按 owner 提供的正式中文名称补齐。"},
}


def normalized_id(value: Any) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value or ""))
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", without_marks.lower())


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def nonempty(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=SOURCE_WORKBOOK_PATH)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.input.resolve()
    output_path = args.output.resolve()
    if workbook_path != SOURCE_WORKBOOK_PATH.resolve():
        raise RuntimeError(f"ITEM_LOCALIZATION_SOURCE_PATH: expected {SOURCE_WORKBOOK_RELATIVE_PATH}, received {workbook_path}")
    if not workbook_path.is_file():
        raise RuntimeError(f"ITEM_LOCALIZATION_SOURCE_MISSING: {SOURCE_WORKBOOK_RELATIVE_PATH}")
    workbook_sha256 = sha256(workbook_path)
    if workbook_path.stat().st_size != SOURCE_WORKBOOK_SIZE or workbook_sha256 != SOURCE_WORKBOOK_SHA256:
        raise RuntimeError(f"ITEM_LOCALIZATION_SOURCE_FINGERPRINT: expected {SOURCE_WORKBOOK_SIZE}/{SOURCE_WORKBOOK_SHA256}, received {workbook_path.stat().st_size}/{workbook_sha256}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_names = set(workbook.sheetnames)
        required = {"道具列表", "道具列表2"}
        if not required.issubset(sheet_names):
            raise RuntimeError(f"ITEM_LOCALIZATION_SHEETS_MISSING: {sorted(required - sheet_names)}")
        primary = workbook["道具列表"]
        headers = [nonempty(value) for value in next(primary.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            zh_column = headers.index("简体中文")
            english_column = headers.index("英文")
        except ValueError as error:
            raise RuntimeError("ITEM_LOCALIZATION_HEADERS_MISSING") from error

        source_rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(primary.iter_rows(min_row=2, values_only=True), start=2):
            english = nonempty(row[english_column]) if english_column < len(row) else None
            chinese = nonempty(row[zh_column]) if zh_column < len(row) else None
            if english is None:
                continue
            source_rows.append({"sourceRow": row_number, "english": english, "zhHansName": chinese})
        secondary = workbook["道具列表2"]
        secondary_rows = [
            {"sourceRow": row_number, "english": nonempty(row[3]) if len(row) > 3 else None, "zhHansName": nonempty(row[1]) if len(row) > 1 else None}
            for row_number, row in enumerate(secondary.iter_rows(min_row=2, values_only=True), start=2)
        ]
    finally:
        workbook.close()

    runtime_path = workbook_path.parent.parent / "public" / "data" / "items.json"
    registry_path = workbook_path.parent.parent / "data-curated" / "id-registry.json"
    runtime_items = json.loads(runtime_path.read_text(encoding="utf-8"))
    registry = json.loads(registry_path.read_text(encoding="utf-8"))["entities"]
    active_registry = {entry["projectId"]: entry for entry in registry if entry.get("kind") == "item" and entry.get("status") == "active"}
    if len(runtime_items) != 567 or len(active_registry) != 567:
        raise RuntimeError(f"ITEM_LOCALIZATION_LEFT_TABLE_COUNT: runtime={len(runtime_items)} registry={len(active_registry)}")

    by_showdown_id: dict[str, list[dict[str, Any]]] = {}
    for row in source_rows:
        by_showdown_id.setdefault(normalized_id(row["english"]), []).append(row)

    records: list[dict[str, Any]] = []
    for item in runtime_items:
        stable_id = item["itemId"]
        registry_entry = active_registry.get(stable_id)
        if registry_entry is None:
            raise RuntimeError(f"ITEM_LOCALIZATION_REGISTRY_MISSING: {stable_id}")
        showdown_id = registry_entry["showdownId"]
        canonical_name = item["canonicalName"]
        if normalized_id(canonical_name) != showdown_id:
            raise RuntimeError(f"ITEM_LOCALIZATION_IDENTITY_MISMATCH: {stable_id}")

        candidates = by_showdown_id.get(showdown_id, [])
        distinct_names = list(dict.fromkeys(row["zhHansName"] for row in candidates if row["zhHansName"]))
        override = OWNER_OVERRIDES.get(showdown_id)
        if override is not None:
            mapping_class = "owner-override"
            zh_name = override["zhHansName"]
            selected = next((row for row in candidates if row["zhHansName"] == zh_name), None)
        elif len(distinct_names) == 1:
            mapping_class = "automatic"
            zh_name = distinct_names[0]
            selected = next(row for row in candidates if row["zhHansName"] == zh_name)
        else:
            raise RuntimeError(f"ITEM_LOCALIZATION_AMBIGUOUS_OR_MISSING: {stable_id}:{canonical_name}")

        records.append({
            "stableId": stable_id,
            "showdownId": showdown_id,
            "canonicalEnglishName": canonical_name,
            "zhHansName": zh_name,
            "mappingClass": mapping_class,
            "sourceProvenance": {
                "sourceId": "52poke-wiki-item-names",
                "sourceName": SOURCE_NAME,
                "sourceUrl": SOURCE_URL,
                "sourceTable": "道具列表",
                "sourceLocator": "manual-list-snapshot",
                "sourceRow": selected["sourceRow"] if selected is not None else None,
                "sourceEnglishName": selected["english"] if selected is not None else None,
                "copiedAt": COPIED_AT,
                "ownerOverride": mapping_class == "owner-override",
            },
        })

    records.sort(key=lambda record: record["stableId"])
    secondary_by_showdown_id: dict[str, set[str]] = {}
    for row in secondary_rows:
        if row["english"] is not None and row["zhHansName"] is not None:
            secondary_by_showdown_id.setdefault(normalized_id(row["english"]), set()).add(row["zhHansName"])
    secondary_comparisons = [
        {
            "stableId": record["stableId"],
            "canonicalEnglishName": record["canonicalEnglishName"],
            "formalZhHansName": record["zhHansName"],
            "secondaryZhHansNames": sorted(secondary_by_showdown_id[record["showdownId"]]),
        }
        for record in records
        if record["showdownId"] in secondary_by_showdown_id
    ]
    secondary_mismatches = [
        comparison for comparison in secondary_comparisons
        if len(comparison["secondaryZhHansNames"]) != 1 or comparison["formalZhHansName"] not in comparison["secondaryZhHansNames"]
    ]
    overrides = []
    for showdown_id, override in OWNER_OVERRIDES.items():
        record = next(record for record in records if record["showdownId"] == showdown_id)
        overrides.append({
            "stableId": record["stableId"],
            "showdownId": showdown_id,
            "canonicalEnglishName": record["canonicalEnglishName"],
            "zhHansName": override["zhHansName"],
            "reason": override["reason"],
            "sourceCandidates": [
                {"sourceRow": row["sourceRow"], "sourceEnglishName": row["english"], "zhHansName": row["zhHansName"]}
                for row in by_showdown_id.get(showdown_id, [])
            ],
        })
    overrides.sort(key=lambda record: record["stableId"])

    result = {
        "schemaVersion": 1,
        "scope": "active Item registry only",
        "source": {
            "sourceId": "52poke-wiki-item-names",
            "sourceName": SOURCE_NAME,
            "sourceUrl": SOURCE_URL,
            "copiedAt": COPIED_AT,
            "sourceTable": "道具列表",
            "mappingMethod": "existing 567 active Item registry -> English canonical / Showdown ID -> 52Poké English column -> 简体中文 column -> explicit owner override",
            "provenanceMode": "manual list snapshot / manually copied source; no per-Item page or revision URL is asserted",
            "license": {
                "name": "CC BY-NC-SA 3.0",
                "url": LICENSE_URL,
                "attributionBoundary": "short Item-name localization only",
            },
            "transportEvidence": {
                "sourceWorkbook": SOURCE_WORKBOOK_RELATIVE_PATH,
                "sourceWorkbookSize": SOURCE_WORKBOOK_SIZE,
                "sourceWorkbookSha256": workbook_sha256,
                "sourceRole": "independent 52Poké manual localization workbook",
                "primarySheet": "道具列表",
                "secondarySheet": "道具列表2",
            },
            "secondarySource": {
                "sheet": "道具列表2",
                "role": "cross-check only; never overrides formal names",
                "descriptionIntegration": "excluded from this production integration",
                "crossCheck": {
                    "dataRows": len(secondary_rows),
                    "comparedActiveItems": len(secondary_comparisons),
                    "agreeingItems": len(secondary_comparisons) - len(secondary_mismatches),
                    "mismatchItems": len(secondary_mismatches),
                    "mismatches": secondary_mismatches,
                },
            },
        },
        "ownerOverrides": overrides,
        "records": records,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(output_path),
        "records": len(records),
        "automatic": sum(record["mappingClass"] == "automatic" for record in records),
        "ownerOverrides": sum(record["mappingClass"] == "owner-override" for record in records),
        "workbookSha256": workbook_sha256,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
